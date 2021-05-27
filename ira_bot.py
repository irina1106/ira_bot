from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook

TOKEN = '1693530992:AAGvNISP986tbRL2xgaIevs75kJRZWUkXrk'
book = load_workbook('database.xlsx')
sheet_1 = book['Лист1']
films_page = book['фильмы']


def main():
    updater = Updater(token=TOKEN)
    dispatcher = updater.dispatcher
    handler = MessageHandler(Filters.all, do_echo)
    start_handler = CommandHandler('start', do_start)
    help_handler = CommandHandler('help', do_help)
    films_handler = MessageHandler(Filters.all, do_films)

    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(help_handler)
    dispatcher.add_handler(films_handler)
    dispatcher.add_handler(handler)
    updater.start_polling()
    updater.idle()


def do_echo(update, context):
    text = update.message.text

    if text == "привет":
        update.message.reply_text(text="привет")
    else:
        update.message.reply_text(text="привет! нажми /start")



def do_start(update, context):
    keyboard = [
        ["комедия", "боевик", "драма", "детектив", "мелодрама"]
    ]
    update.message.reply_text(
        text="выбери жанр и увидишь список фильмов",
        reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True))


def do_help(update, context):
    update.message.reply_text(text="тебе чем-то помочь?")


def get_films(genre):
    reply_text = ""
    for i in range(1, 9):
        reply_text += films_page.cell(row=i, column=genre).value + "\n"
    return reply_text

def do_films(update, context):
    text = update.message.text
    genres = ["комедия", "боевик", "драма", "детектив", "мелодрама"]
    if text == "привет":
        update.message.reply_text(text="привет! нажми /start")
        return

    for i in genres:
        if text == i:
            if text == "комедия":
                update.message.reply_text(text=get_films(1))
            elif text == "боевик":
                update.message.reply_text(text=get_films(2))
            elif text == "драма":
                update.message.reply_text(text=get_films(3))
            elif text == "детектив":
                update.message.reply_text(text=get_films(4))
            elif text == "мелодрама":
                update.message.reply_text(text=get_films(5))
            return



main()