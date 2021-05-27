from openpyxl import load_workbook


book = load_workbook('database.xlsx')
sheet_1 = book['Лист1']
films_page = book['фильмы']

print(book.worksheets)
for i in range(1, 9):
    print(films_page.cell(row=i, column=1).value)
